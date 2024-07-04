
import os
import frappe
import datetime
from datetime import datetime
import math
from frappe.utils.data import add_to_date
from frappe.utils.file_manager import save_file_on_filesystem
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

@frappe.whitelist(allow_guest=True)
def create_xlsx(from_date, to_date):
     
    # Define the path to the Downloads folder
    downloads_folder = os.path.expanduser("~/Downloads")
    
    # Define the base filename
    base_filename = "ESI_Return.xlsx"
    
    # Check if the file already exists
    file_number = 1
    while True:
        # Construct the file path
        if file_number == 1:
            file_path = os.path.join(downloads_folder, base_filename)
        else:
            file_path = os.path.join(downloads_folder, f"ESI_Return({file_number}).xlsx")
        
        # If the file doesn't exist, break the loop
        if not os.path.exists(file_path):
            break
        
        # Increment the file number
        file_number += 1
    
    # Create a new Workbook
    wb = Workbook()
    
    # Get the default worksheet
    first_ws = wb.active
    first_ws.title = "Sheet 1"
    
    # Add content to the first sheet
    # first_ws['A1'] = "This is content for the first sheet."

    wb = Workbook()
    
    # Get the default worksheet
    first_ws = wb.active
    first_ws.title = "Sheet 1"
    
    # Add content to the first sheet
    first_sheet_data = [
        ['IP Number', 'IP Name', 'No of Days for which wages paid/payable during the month',
         'Total Monthly Wages', 'Reason Code', 'Last Working Day']
    ]

    for row in first_sheet_data:
        first_ws.append(row)

    salary_slips = frappe.get_all("Salary Slip", filters={
        "start_date": ["<=", from_date],
        "end_date": [">=", to_date],
        "docstatus": 1
    }, fields=['name', 'esi_number', 'employee_name', 'gross_pay', 'payment_days', 'reason_code', 'custom_relieving_date'])

    results = []

    for slip in salary_slips:
        esi_number = slip.get('esi_number')
        employee_name = slip.get('employee_name')
        payment_days = math.ceil(slip.get('payment_days'))
        gross_pay = math.ceil(slip.get('gross_pay'))
        reason_code = slip.get('reason_code')
        custom_relieving_date = slip.get('custom_relieving_date')

        if custom_relieving_date:
            custom_relieving_date = datetime.strptime(custom_relieving_date, '%Y-%m-%d').strftime('%d/%m/%Y')

        results.append((esi_number, employee_name, payment_days, gross_pay, reason_code, custom_relieving_date))

    # Append data to the first sheet
    for result in results:
        first_ws.append(result)

    first_ws.column_dimensions['A'].width = 20
    first_ws.column_dimensions['B'].width = 30
    first_ws.column_dimensions['C'].width = 20
    first_ws.column_dimensions['D'].width = 20
    first_ws.column_dimensions['E'].width = 20
    first_ws.column_dimensions['F'].width = 30

    # # Apply background color to the first row
    # for cell in first_ws[1]:
    #     cell.fill = PatternFill(start_color="ccffff", end_color="ccffff", fill_type="solid")
        
    # Make 'Reason', 'Code', and 'Note' headings bold and apply background color
    for col in range(1, 7):
        heading = first_ws.cell(row=1, column=col)
        heading.font = Font(bold=True, name='Aharoni', size=11)
        heading.fill = PatternFill(start_color="ccffff", end_color="ccffff", fill_type="solid")
        

    # Apply borders to the table
    rows = len(results) + 1  # +1 for the header row
    cols = len(first_sheet_data[0])
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            cell = first_ws.cell(row=row, column=col)
            cell.border = Border(left=Side(border_style='thin'), 
                                 right=Side(border_style='thin'), 
                                 top=Side(border_style='thin'), 
                                 bottom=Side(border_style='thin'))
            
     # Center align columns C, D, and E
    for col in range(3, 6):
        for cell in first_ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                c.alignment = Alignment(horizontal='center')       
            

    # Create a new worksheet
    second_ws = wb.create_sheet("Instructions & Reason Codes")
    
    # Add content to the second sheet
    second_ws['A1'] = " "

    # Create a table in the second sheet
    data = [
        ['Reason', 'Code', 'Note'],
        ['Without Reason', 0, 'Leave last working day as blank'],
        ['On Leave', 1, 'Leave last working day as blank'],
        ['Left Service', 2, 'Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period'],
        ['Retired', 3, 'Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period'],
        ['Out of Coverage' , 4, 'Please provide last working day (dd/mm/yyyy). IP will not appear from next contribution period. This option is valid only if Wage Period is April/October. In case any other month then IP will continue to appear in the list'],
        ['Expired' , 5, 'Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period'],
        ['Non Implemented area' , 6, 'Please provide last working day (dd/mm/yyyy).'],
        ['Compliance by Immediate Employer' , 7, 'Leave last working day as blank'],
        ['Suspension of work' , 8, 'Leave last working day as blank'],
        ['Strike/Lockout' , 9, 'Leave last working day as blank'],
        ['Retrenchment' , 10, 'Please provide last working day (dd/mm/yyyy). IP will not appear from next wage period'],
        ['No Work' , 11, 'Leave last working day as blank'],
        ['Doesnt Belong To This Employer' , 12, 'Leave last working day as blank'],
        ['Duplicate IP' , 13, 'Leave last working day as blank']
    ]
    for row in data:
        second_ws.append(row)

    # Create a table
    table = openpyxl.worksheet.table.Table(displayName="Table1", ref="A1:C4")

    # Add a style
    style = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    second_ws.add_table(table)

    # Add borders to the table
    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    for row in second_ws.iter_rows(min_row=2, max_row=16, min_col=1, max_col=3):
        for cell in row:
            cell.border = border
    
    
    
    # Center align the 'Code' column
    for cell in second_ws['B']:
        cell.alignment = Alignment(horizontal='center')

    # Center align the 'Note' heading
    second_ws['C2'].alignment = Alignment(horizontal='center')
    
    # Make 'Reason', 'Code', and 'Note' headings bold and apply background color
    for col in range(1, 4):
        heading_cell = second_ws.cell(row=2, column=col)
        heading_cell.font = Font(bold=True, name='Arial', size=10)
        heading_cell.fill = PatternFill(start_color="99ccff", end_color="99ccff", fill_type="solid")
        
    # Apply font style 'Arial' with size 10pt to all data in the table
    for row in second_ws.iter_rows(min_row=3, max_row=16, min_col=1, max_col=3):
        for cell in row:
            cell.font = Font(name='Arial', size=10)
            
    second_ws.column_dimensions['C'].width = 90
    second_ws.column_dimensions['A'].width = 20
    
    # Set the line "Click Here to Go back to Data Entry Page" in the 18th row
    link_text = "Click Here to Go back to Data Entry Page"
    link_cell = second_ws.cell(row=18, column=1)
    link_cell.value = link_text
    link_font = Font(color="FF0000", name="Arial", size=16, underline="single")
    link_cell.font = link_font
    
    # Set the instructions in the 20th row
    instruction_text = "Instructions to fill in the excel file:"
    instruction_cell = second_ws.cell(row=20, column=1)
    instruction_cell.value = instruction_text
    instruction_font = Font(name="Arial", size=14, bold=True)
    instruction_cell.font = instruction_font
    
    # Add instructions in rows 21 to 35
    instructions = [
        "1. Enter the IP number, IP name, No. of Days, Total Monthly Wages, Reason for 0 wages(If Wages ‘0’) & Last Working Day(only if employee has left service, Retired, Out of coverage, Expired, Non-Implemented area or Retrenchment. For other reasons, last working day must be left BLANK).",
        "2. Number of days must me a whole number. Fractions should be rounded up to next higher whole number/integer",
        "3. Excel sheet upload will lead to successful transaction only when all the Employees’ (who are currently mapped in the system) details are entered perfectly in the excel sheet",
        "4. Reasons are to be assigned numeric code and date has to be provided as mentioned in the table above",
        "5. Once 0 wages given and last working day is mentioned as in reason codes (2,3,4,5,10) IP will be removed from the employer’s record. Subsequent months will not have this IP listed under the employer. Last working day should be mentioned only if 'Number of days wages paid/payable' is '0'.",
        "6. In case IP has worked for part of the month(i.e. atleast 1 day wage is paid/payable) and left in between of the month, then last working day shouldn’t be mentioned.",
        "7. Calculations – IP Contribution and Employer contribution calculation will be automatically done by the system",
        "8. Date column format is dd/mm/yyyy or dd-mm-yyyy. Pad single digit dates with 0. Eg:- 2/5/2010 or 2-May-2010 is NOT acceptable. Correct format is 02/05/2010 or 02-05-2010",
        "9. Excel file should be saved in .xls format (Excel 97-2003)",
        "10. Note that all the column including date column should be in ‘Text’ format",
        "10a. To convert all columns to text,",
        "    a. Select column A; Click Data in Menu Bar on top; Select Text to Columns ; Click Next (keep default selection of Delimited); Click Next (keep default selection of Tab); Select TEXT; Click FINISH. Excel 97 – 2003 as well have TEXT to COLUMN conversion facility",
        "    b. Repeat the above step for each of the 6 columns. (Columns A – F )",
        "10b. Another method that can be used to text conversion is – copy the column with data and paste it in NOTEPAD. Select the column (in excel) and convert to text. Copy the data back from notepad to excel",
        "11. If problem continues while upload, download a fresh template by clicking 'Sample MC Excel Template'. Then copy the data area from Step 8a.a – eg: copy Cell A2 to F8 (if there is data in 8 rows); Paste it in cell A2 in the fresh template. Upload it"
    ]
    instruction_font = Font(name="Arial", size=10)
    for index, instruction in enumerate(instructions, start=21):
        instruction_cell = second_ws.cell(row=index, column=1)
        instruction_cell.value = instruction
        instruction_cell.font = instruction_font
        
    # Add the note in the 36th row
    note_text = "Note: Kindly turn OFF 'POP UP BLOCKER' if it is ON in your browser. Follow the steps given to turn off pop up blocker. This is required to upload Monthly contribution, view or print Challan / TIC after uploading the excel"
    note_cell = second_ws.cell(row=37, column=1)
    note_cell.value = note_text

    # Apply formatting to the note
    note_font = Font(name="Calibri", size=11)
    note_cell.font = note_font
    
    # Add the Firefox instruction in the 38th row
    firefox_instruction_text = "              1. Mozilla Firefox 3.5.11: From Menu Bar, select Tools -> Options -> Content -> Uncheck (remove tick mark) 'Block Popup Windows'. Click OK"
    firefox_instruction_cell = second_ws.cell(row=38, column=1)
    firefox_instruction_cell.value = firefox_instruction_text
    firefox_instruction_font = Font(name="Calibri", size=11)
    firefox_instruction_cell.font = firefox_instruction_font

    # Add the IE instruction in the 39th row
    ie_instruction_text = "              2. IE 7.0:   From Menu Bar, select Tools -> Pop up Blocker -> Turn Off Pop up Blocker"
    ie_instruction_cell = second_ws.cell(row=39, column=1)
    ie_instruction_cell.value = ie_instruction_text
    ie_instruction_cell.font = firefox_instruction_font

    # Save the workbook to the Downloads folder
    wb.save(file_path)
    
    # Save the file to the filesystem
    save_file_on_filesystem(file_path, content=open(file_path, "rb").read())
    
    return file_path


@frappe.whitelist(allow_guest=True)
def mark_attendance(date, shift):
    success_message_printed = False

    checkin_records = frappe.db.get_all(
        "Employee Checkin",
        filters={
            "shift": shift,
            "date": date
        },
        fields=["employee", "name", "date", "log_type"],
        order_by="date"
    )

    # if not checkin_records:
    #     date_obj = datetime.strptime(date, '%Y-%m-%d')
    #     formatted_date = date_obj.strftime('%d-%m-%Y')
    #     frappe.msgprint(f"No Checkin Records found for the date {formatted_date}")
    # else:
    result_dict = {}

    for record in checkin_records:
        employee_id = record["employee"]
        checkin_date = record["date"]

        if employee_id not in result_dict:
            result_dict[employee_id] = {}

        if checkin_date not in result_dict[employee_id]:
            result_dict[employee_id][checkin_date] = []

        result_dict[employee_id][checkin_date].append({
            "name": record["name"],
            "log_type": record["log_type"]
        })

    first_chkin = None
    last_chkout = None

    for employee_id, dates in result_dict.items():
        for checkin_date, logs in dates.items():
            first_chkin = None
            last_chkout = None

            for log in logs:
                name = log['name']
                log_type = log['log_type']

                if log_type == "IN" and first_chkin is None:
                    first_chkin = name

                if log_type == "OUT":
                    last_chkout = name

            if first_chkin and last_chkout:
                exits_atte = frappe.db.get_value('Attendance', {'employee': employee_id, 'attendance_date': checkin_date, 'docstatus': 1}, ['name'])
                if not exits_atte:
                    
                    chkin_datetime = frappe.db.get_value('Employee Checkin', first_chkin, 'time')
                    chkout_datetime = frappe.db.get_value('Employee Checkin', last_chkout, 'time')

                    chkin_time = frappe.utils.get_time(chkin_datetime)
                    chkout_time = frappe.utils.get_time(chkout_datetime)

                    attendance = frappe.new_doc("Attendance")
                    attendance.employee = employee_id
                    attendance.attendance_date = checkin_date
                    attendance.shift = shift
                    attendance.in_time = chkin_datetime
                    attendance.out_time = chkout_datetime
                    attendance.check_in_time = chkin_time
                    attendance.check_out_time = chkout_time
                    attendance.custom_employee_checkin = first_chkin
                    attendance.custom_employee_checkout = last_chkout
                    attendance.status = "Present"

                    attendance.insert(ignore_permissions=True)
                    attendance.submit()
                    frappe.db.commit()

                    if not success_message_printed:
                        frappe.msgprint("Attendance is Marked Successfully")
                        success_message_printed = True
                else:
                    attendance_link = frappe.utils.get_link_to_form("Attendance", exits_atte)
                    frappe.msgprint(f"Attendance already marked of Employee:{employee_id} for date {checkin_date}: {attendance_link}")

            elif first_chkin and not last_chkout:
                exits_atte = frappe.db.get_value('Attendance', {'employee': employee_id, 'attendance_date': checkin_date, 'docstatus': 1}, ['name'])
                if not exits_atte:
                    chkin_datetime = frappe.db.get_value('Employee Checkin', first_chkin, 'time')
                    chkin_time = frappe.utils.get_time(chkin_datetime)

                    attendance = frappe.new_doc("Attendance")
                    attendance.employee = employee_id
                    attendance.attendance_date = checkin_date
                    attendance.shift = shift
                    attendance.in_time = chkin_datetime
                    attendance.check_in_time = chkin_time
                    attendance.custom_employee_checkin = first_chkin
                    attendance.status = "Present"
                    attendance.custom_remarks = "No OutPunch"

                    attendance.insert(ignore_permissions=True)
                    attendance.submit()
                    frappe.db.commit()

                    if not success_message_printed:
                        frappe.msgprint("Attendance is Marked Successfully")
                        success_message_printed = True
                else:
                    attendance_link = frappe.utils.get_link_to_form("Attendance", exits_atte)
                    frappe.msgprint(f"Attendance already marked of Employee:{employee_id} for date {checkin_date}: {attendance_link}")



                  
@frappe.whitelist(allow_guest=True)
def set_attendance_date():
    
    yesterday_date = add_to_date(datetime.now(), days=-1)
    date = yesterday_date.strftime('%Y-%m-%d')

    shift_types = frappe.get_all("Shift Type", filters={'enable_auto_attendance':1},fields=['name'])
    if shift_types:
        for shifts in shift_types:
            shift = shifts.name

            mark_attendance(date, shift)


@frappe.whitelist(allow_guest=True)
def get_gratuity(employee):
    today = datetime.date.today()
    salary_slips = frappe.get_all("Salary Slip", filters={"employee": employee, "docstatus": 1}, fields=['name', 'employee', 'start_date', 'end_date'])
    
    if salary_slips:
        salary_slips.sort(key=lambda slip: abs((slip['end_date'] - today).days))
        get_basic_salary = frappe.get_all('Salary Detail', filters={
            'parent': ['in', [ss['name'] for ss in salary_slips]],
            'salary_component': 'Basic'  
        }, fields=['amount'])
        return get_basic_salary[0]
        # return salary_slips[0], get_basic_salary[0]
    else:
        return None


@frappe.whitelist(allow_guest=True)
def get_employee_bonus(employee, from_date, to_date):
    if not (employee and from_date and to_date):
        return None  

    from_date = datetime.datetime.strptime(from_date, '%Y-%m-%d')
    to_date = datetime.datetime.strptime(to_date, '%Y-%m-%d')

    salary_slips = frappe.get_all("Salary Slip", filters={
        "docstatus": 1,
        "employee": employee,
        "start_date": (">=", from_date),
        "end_date": ("<=", to_date)
    }, fields=['name', 'start_date', 'end_date'])

    if not salary_slips:
        return None  

    get_salary_data = frappe.get_all('Salary Detail', filters={
        'parent': ['in', [ss['name'] for ss in salary_slips]],
        'salary_component': 'Basic'  
    }, fields=['amount'])
    
    return salary_slips, get_salary_data


@frappe.whitelist(allow_guest=True)
def generate_txt(selected_date):
    salary_slips = frappe.get_all("Salary Slip", filters={
        "start_date": ["<=", selected_date],
        "end_date": [">=", selected_date],
        "docstatus": 1
    }, fields=['name', 'employee', 'uan', 'gross_pay', 'leave_without_pay', 'absent_days'])
    # return salary_slips
    output_data = []
    # stored_basic_amount = None
    
    if salary_slips:
        for salary_slip in salary_slips:
            leave_without_pay = salary_slip.leave_without_pay
            absent_days = salary_slip.absent_days
            ncp_days = leave_without_pay + absent_days
            if ncp_days == 0.0:
                ncp_days = 0
            uan = salary_slip.uan
            gross_pay = salary_slip.gross_pay
            employee_name = frappe.get_value("Employee", salary_slip.employee, "employee_name")
            salary_detail_data = frappe.get_all('Salary Detail', filters={
                'parent': salary_slip.name, 
                'salary_component': ['in', ['Basic', 'Provident Fund']]
            }, fields=['salary_component', 'amount'])

            basic_amount = None
            pf_amount = None

            for detail in salary_detail_data:
                if detail['salary_component'] == 'Basic':
                    basic_amount = format_amount(detail['amount'])
                    eps = basic_amount * (8.33 / 100)
                    # print("EPS:", eps)
                    # stored_basic_amount = basic_amount
                elif detail['salary_component'] == 'Provident Fund':
                    pf_amount = format_amount(detail['amount'])
                    # print("pf:", pf_amount)

            diff = pf_amount - eps
            # print("diff:", diff)
            gross_pay_formatted = format_amount(gross_pay)
            ncp_days_formatted = format_amount(ncp_days)

            employee_output = f"{uan}#~#{employee_name}#~#{math.ceil(gross_pay_formatted)}#~#{math.ceil(basic_amount)}#~#{math.ceil(basic_amount)}#~#{math.ceil(basic_amount)}#~#{math.ceil(pf_amount)}#~#{math.ceil(eps)}#~#{math.ceil(diff)}#~#{math.ceil(ncp_days_formatted)}#~#0"

            output_data.append(employee_output)

        final_output = "\n".join(output_data)
        print(final_output)
        txt_file_path = "PF_ECR.txt"
        save_file_on_filesystem(txt_file_path, content=final_output, is_private=0)
        return txt_file_path
            
    else: 
        frappe.msgprint("Salary slip does not exist for given dates.")

def format_amount(amount):
    if amount == int(amount):
        return float(amount)
    else:
        return float(amount)


@frappe.whitelist(allow_guest=True)
def generate_leave_allocation():
    # return "Hello"
    current_date = datetime.datetime.now().date().strftime('%Y-%m-%d')
    
    l_type = frappe.db.get_list('Leave Type', fields="name", filters={"is_privilege_leave":1})
    # return l_type 

    for leaves in l_type:
        leave_type = leaves.name
    
        all_emp = frappe.db.get_list('Leave Allocation', filters={'leave_type': leave_type}, fields=['employee_name'])
        
        for emp in all_emp:
            employee_name = emp.employee_name
            last_doc = frappe.get_last_doc("Leave Allocation", filters={"employee_name": employee_name, "docstatus": 1})
            
            if last_doc:
                to_date = last_doc.to_date
                next_date = to_date + datetime.timedelta(days=1)
                specific_fields = {
                    "emp": last_doc.employee,
                    "emp_name": last_doc.employee_name,
                    "leave_type": last_doc.leave_type,
                    "to_date": last_doc.to_date,
                    "next_date": next_date
                }
            else:
                specific_fields = None
                to_date = None 
            
            if to_date and to_date.strftime('%Y-%m-%d') == current_date:  
                new_leave_allocation = next_date + datetime.timedelta(days=19)
                leave_period = frappe.get_all("Leave Period", filters={"from_date": ("<=", next_date), "to_date": (">=", next_date)}, fields=["name", "from_date", "to_date"])
                
                if leave_period:
                    leave_period = leave_period[0]
                    
                    from_date_leave_period = leave_period.get("from_date")
                    to_date_leave_period = leave_period.get("to_date")
                    
                    count_leave_allocations = frappe.get_all("Leave Allocation", filters={"employee": specific_fields["emp"], "from_date": (">=", from_date_leave_period), "to_date": ("<=", to_date_leave_period)})
                    count = len(count_leave_allocations)
                    
                    if count < 15 :
                        existing_allocation = frappe.get_all("Leave Allocation", filters={"employee": specific_fields["emp"], "from_date": next_date, "to_date": new_leave_allocation})
                        
                        if not existing_allocation:
                            if new_leave_allocation <= to_date_leave_period:
                                new_allocation = frappe.new_doc("Leave Allocation")
                                new_allocation.employee = specific_fields["emp"]
                                new_allocation.leave_type = specific_fields["leave_type"]
                                new_allocation.from_date = next_date
                                new_allocation.to_date = new_leave_allocation
                                new_allocation.new_leaves_allocated = "1"
                                new_allocation.carry_forward = 1
                                new_allocation.insert(ignore_permissions=True)
                                frappe.db.commit()
                            else:
                                new_allocation = frappe.new_doc("Leave Allocation")
                                new_allocation.employee = specific_fields["emp"]
                                new_allocation.leave_type = specific_fields["leave_type"]
                                new_allocation.from_date = next_date
                                new_allocation.to_date = to_date_leave_period
                                new_allocation.new_leaves_allocated = "1"
                                new_allocation.carry_forward = 1
                                new_allocation.insert(ignore_permissions=True)
                                frappe.db.commit()
                        else:
                            None
                            # frappe.msgprint("Leave allocation already exists for the given period")
                    else: 
                        None
                        # frappe.msgprint("Can't allocate leaves more than 15")
                else:
                    None
                    # frappe.msgprint("Create Leave Period for the current month")
            else:
                return "NO"

    #     return "yesssssssssss"  
    # return "yes" 


